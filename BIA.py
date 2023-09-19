from tkinter import *
from tkinter import filedialog, messagebox, ttk, PhotoImage
import openpyxl
from openpyxl.utils import get_column_letter



# will process the data to select the most suitable student for each project
def Processar(arquivo):
    try:
        # setting the location of the selected file and the tabs where the data will be located
        livro = openpyxl.load_workbook(f'{arquivo}')

        pagina_aluno = livro['Indicações dos bolsistas']
        pagina_professor = livro['Indicação dos orientadores']

        alunos = []
        professores = []

        # Counting the number of lines
        linhas_preenchidasBolsistas = 0
        for celula in pagina_aluno["A"]:
            if (celula.value != None):
                linhas_preenchidasBolsistas += 1
        linhas_preenchidasOrientadores = 0
        for celula in pagina_professor["A"]:
            if (celula.value != None):
                linhas_preenchidasOrientadores += 1

        # Looking for the data that will be used in the excel cells
        nome = ""
        opcao1 = ""
        opcao2 = ""
        opcao3 = ""
        opcao4 = ""
        opcao5 = ""
        pAlocado = ""

        for coluna, celula in enumerate(pagina_aluno["2"], start=1):
            if celula.value == "Nome":
                nome = get_column_letter(coluna)
            if celula.value == "1° opção":
                opcao1 = get_column_letter(coluna)
            if celula.value == "2° opção":
                opcao2 = get_column_letter(coluna)
            if celula.value == "3° opção":
                opcao3 = get_column_letter(coluna)
            if celula.value == "4° opção":
                opcao4 = get_column_letter(coluna)
            if celula.value == "5° opção":
                opcao5 = get_column_letter(coluna)
            if celula.value != None and "Projeto alocado" in celula.value:
                pAlocado = get_column_letter(coluna)

        titulo = ""
        vagas = ""
        responsavel = ""
        bolsistaAlocado = ""
        opcaoOri1 = ""
        opcaoOri2 = ""
        opcaoOri3 = ""
        opcaoOri4 = ""
        opcaoOri5 = ""
        for coluna, celula in enumerate(pagina_professor["1"], start=1):
            if celula.value != None and "TÍTULO DO PROJETO" in celula.value:
                titulo = get_column_letter(coluna)
            if celula.value != None and "RESPONSÁVEL" in celula.value:
                responsavel = get_column_letter(coluna)
            if celula.value != None and "VAGAS" in celula.value:
                vagas = get_column_letter(coluna)
            if celula.value != None and "BOLSISTA ALOCADO(A)" in celula.value:
                bolsistaAlocado = get_column_letter(coluna)
            if celula.value != None and "1° opção" in celula.value:
                opcaoOri1 = get_column_letter(coluna)
            if celula.value != None and "2° opção" in celula.value:
                opcaoOri2 = get_column_letter(coluna)
            if celula.value != None and "3° opção" in celula.value:
                opcaoOri3 = get_column_letter(coluna)
            if celula.value != None and "4° opção" in celula.value:
                opcaoOri4 = get_column_letter(coluna)
            if celula.value != None and "5° opção" in celula.value:
                opcaoOri5 = get_column_letter(coluna)


        # Storing excel data in a dictionary
        for linha in range(3, linhas_preenchidasBolsistas+2):
            aluno = {
                "nomeAluno": pagina_aluno[f'{nome}{linha}'].value,
                "projetos": [pagina_aluno[f'{opcao1}{linha}'].value, pagina_aluno[f'{opcao2}{linha}'].value, pagina_aluno[f'{opcao3}{linha}'].value, pagina_aluno[f'{opcao4}{linha}'].value, pagina_aluno[f'{opcao5}{linha}'].value],
                "projetoAlocado": "",
                "orientador": False,
                "linha": linha,
            }
            alunos.append(aluno)

        for linha in range(2, linhas_preenchidasOrientadores+2):
            projetos = {
                "orientador": pagina_professor[f'{responsavel}{linha}'].value,
                "projeto": pagina_professor[f'{titulo}{linha}'].value,
                "vagas": pagina_professor[f'{vagas}{linha}'].value,
                "bolsistas": [pagina_professor[f'{opcaoOri1}{linha}'].value, pagina_professor[f'{opcaoOri2}{linha}'].value, pagina_professor[f'{opcaoOri3}{linha}'].value, pagina_professor[f'{opcaoOri4}{linha}'].value, pagina_professor[f'{opcaoOri5}{linha}'].value],
                "bolsistasEscolhido": [],
                "linha": linha
            }
            professores.append(projetos)


        # Will process the data stored in the dictionary to select the student for each project
        for a in alunos:
            for c in range(5):
                for p in professores:
                    projetoEscolhidoBolsista = False
                    if (a["orientador"] == False and a["projetos"][c]!=None and (a["projetos"][c].split('-')[-1].strip().replace(".", "") == p["projeto"].split('-')[-1].strip().replace(".", ""))):
                        projetoEscolhidoBolsista = True
                        if (a["nomeAluno"] in p["bolsistas"] and len(p["bolsistasEscolhido"]) < p["vagas"]):
                            print(f"Student {a['nomeAluno']} foi was chosen for the teacher's project {p['orientador']}")
                            a["orientador"] = True
                            a["projetoAlocado"] = a["projetos"][c]
                            pagina_aluno[f'{pAlocado}{a["linha"]}'] = p["projeto"]
                            p["bolsistasEscolhido"].append(a["nomeAluno"])
                    elif (a["orientador"] == True and a["projetoAlocado"]!="" and a["projetos"][c]!=None and (a["projetos"][c].split('-')[-1].strip().strip().replace(".", "") == p["projeto"].split('-')[-1].strip().replace(".", "")) and a["projetos"].index(a["projetos"][c]) < a["projetos"].index(a["projetoAlocado"])):
                        a["projetoAlocado"] = p["projeto"]
                        pagina_aluno[f'{pAlocado}{a["linha"]}'] = p["projeto"]
                        p["bolsistasEscolhido"].append(a["nomeAluno"])

                    if (a["orientador"] == False and (None in p["bolsistas"]) and len(p["bolsistasEscolhido"]) < p["vagas"] and projetoEscolhidoBolsista==True):  # Preciso corrigir aqui alterar o algorítmo para que compare a primeira opção do aluno com todos os projetos
                        print(f"Student {a['nomeAluno']} foi was chosen for the teacher's project {p['orientador']}")
                        a["orientador"] = True
                        pagina_aluno[f'{pAlocado}{a["linha"]}'] = p["projeto"]
                        p["bolsistasEscolhido"].append(a["nomeAluno"])

                    if (len(p["bolsistasEscolhido"])<= p["vagas"] ):
                        aux=''
                        for z in range(len(p["bolsistasEscolhido"])):
                            if (z == p["vagas"]):
                                aux += p["bolsistasEscolhido"][z]
                            else:
                                aux += p["bolsistasEscolhido"][z] + '\n'

                        pagina_professor[f'{bolsistaAlocado}{p["linha"]}'] = aux



            if (a["orientador"] == False):
                print(f'{a["nomeAluno"]} was not chosen by a supervisor')
                pagina_aluno[f'{pAlocado}{a["linha"]}'] = f'There was no mutual indication, analyze manually'

        livro.save(f'{saveTexto.get()}/{diretorio["text"].replace(".xlsx", "")}_preenchida.xlsx')

        messagebox.showinfo(f"Notice", f"Spreadsheet '{diretorio['text'].replace('.xlsx', '')}_preenchida' successfully created!")
        print(professores)

        for c in professores:
            if (len(c["bolsistasEscolhido"])>1):
                for z in range(len(c["bolsistasEscolhido"])):
                    listaAtual = [c["bolsistasEscolhido"][z], c["orientador"], c["projeto"]]
                    tree.insert("", END, values=listaAtual)
            elif (len(c["bolsistasEscolhido"])==1):
                listaAtual = [c["bolsistasEscolhido"][0], c["orientador"], c["projeto"]]
                tree.insert("", END, values=listaAtual)
            else:
                listaAtual = ["", c["orientador"], ""]
                tree.insert("", END, values=listaAtual)
    except Exception as e:
        if varTexto.get() == "":
            messagebox.showerror("Erro", "Please select a worksheet.")
        else:
            messagebox.showerror("Erro", f"Error: inappropriate spreadsheet. Check the correct position of the data via the Help button")


# --------- Functions used in graphical interfaces ---------

# Function used when clicking the open file button
def openFile():
    file = filedialog.askopenfile(mode="r", title="Select a file", filetypes=(("Excel Files","*.xlsx"),))

    if (file!= None):
        diretorioBasename = file.name.split("/")[-1]   # mostrar o nome do arquivo
        diretorio["text"] = diretorioBasename
        varTexto.set(file.name)
    diretorio_save["text"] = file.name.replace("diretorioBasename","")
    saveTexto.set(file.name.replace(diretorioBasename,""))
# Function used when clicking the save file button
def saveFile():
    file = filedialog.askdirectory(mustexist=True)
    diretorio_save["text"] = file
    saveTexto.set(file)
    print(file)
# Function used when clicking the help button (will create a new window with two tabs)
def help():
    janela_help = Toplevel(janela)
    janela_help.title("Ajuda")
    janela_help.geometry("1250x350")
    janela_help.iconbitmap("assets/student.ico")


    tabControl = ttk.Notebook(janela_help)
    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)

    tabControl.add(tab1, text='Indicações dos bolsistas')
    tabControl.add(tab2, text='Indicações dos orientadores')
    tabControl.pack(expand=1, fill="both")

    titulo_descricao = Label(tab1, text="Create a 'Scholarship Recommendations' tab where you will paste the student's data. They need to be on the line indicated in the image, as well as have the exact names in the cells above the data: (detail: students need to be in ascending order by classification)", font=("Ivy 13"), anchor=NE, wraplength=880, pady=30)
    titulo_descricao.pack()

    indicacoesDosBolsistas = PhotoImage(file="assets/IndicaçõesDosBolsistas.png")
    indicacoesDosBolsistas = indicacoesDosBolsistas.subsample(1,1)
    image_label = Label(tab1, image=indicacoesDosBolsistas, pady=20)
    image_label.pack()

    titulo_descricao2 = Label(tab2, text="Create a tab 'Indication of advisors' where you will paste the advisor's data. They must be on the line indicated in the image, as well as have the exact names in the cells above the data:", font=("Ivy 13"), anchor=NE, wraplength=880, pady=30)
    titulo_descricao2.pack()

    indicacaoDosDrientadores = PhotoImage(file="assets/IndicaçãoDosOrientadores.png")
    indicacaoDosDrientadores = indicacaoDosDrientadores.subsample(1,1)
    image_label2 = Label(tab2, image=indicacaoDosDrientadores, pady=20)
    image_label2.pack()

    janela_help.mainloop()

# -------- Start of creating the graphical interface ----------

# Window settings
janela = Tk()
janela.title("BIA - Vacancy allocations")
janela.geometry("600x550")
janela.iconbitmap("assets/student.ico")
varTexto = StringVar()
saveTexto = StringVar()

frame1 = Frame(janela, width=600, height=250, padx=10, pady=10)
frame1.grid(row=0, column=0, sticky="EW")
frame2 = Frame(janela, width=600, height=250, padx=10, pady=10)
frame2.grid(row=1, column=0, sticky="EW")

# Title
titulo = Label(frame1, text="BIA - Vacancy Allocations", font=("Ivy 25"), anchor=NE)
titulo.place(x=5, y=5)

linha = Label(frame1, text='', width=368,anchor=NW, font=("Ivy 1"), bg="#1e3743")
linha.place(x=8, y=50)

# Help button
help_button = Button(frame1, text="Help", width=5, anchor=NE,command=help)
help_button.place(x=530, y=10)

# Upload file
texto_abrir_arquivo = Label(frame1, text="Select the file", font=("Ivy 11"), anchor=NW)
texto_abrir_arquivo.place(x=5, y=80)
botao_dialog = Button(frame1, width=5, height=1,text="Open", fg="white", bg="#1e3743",command=openFile)
botao_dialog.place(x=150, y=78)

diretorio = Label(frame1, text="", font=("Ivy 11"), anchor=NW)
diretorio.place(x=205, y=80)

# Save file
texto_abrir_arquivo = Label(frame1, text="Save location", font=("Ivy 11"), anchor=NW)
texto_abrir_arquivo.place(x=5, y=120)

botao_save = Button(frame1, width=5, height=1,text="Open", fg="white", bg="#1e3743",command=saveFile)
botao_save.place(x=150, y=118)

diretorio_save = Label(frame1, text="", font=("Ivy 11"), anchor=NW)
diretorio_save.place(x=205, y=120)


# Process file
botao = Button(frame1, text="Start", width=20, bg="#1e3743", fg="white",command=lambda: Processar(varTexto.get()))     # create button
botao.place(x=380, y=190)


# -------------- Organizing results in a tree view -------------------------


tree = ttk.Treeview(frame2, selectmode="browse", column=("column1", "column2", "column3"), show="headings")

tree.column("column1", width=150, minwidth=50, stretch=NO)
tree.heading("#1", text="STUDENT")

tree.column("column2", width=150, minwidth=50, stretch=NO)
tree.heading("#2", text="ADVISOR")

tree.column("column3", width=280, minwidth=50, stretch=NO)
tree.heading("#3", text="PROJECT")

tree.grid(row=0, column=0)

elementos = ["Jaime", 32, "Pedro de Queirós"]
# tree.insert("", END, values=elementos)

# Footer
footer = Label(frame2, text="Created by Jaime Jaysi", font=("Ivy 8 bold"), anchor=CENTER)
footer.grid(row=1, column=0, pady=(18, 0), sticky="nsew")

# -------------------------------------------------------------------------------------


janela.mainloop() # will create a loop so that the window stays open